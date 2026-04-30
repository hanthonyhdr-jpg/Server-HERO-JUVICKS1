import os
import uuid
from datetime import datetime
import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from functools import wraps
from werkzeug.utils import secure_filename
import sys
import os
import importlib.util
# Carregamento robusto para evitar erro de PATH no executável
try:
    # 1. Tenta importar como módulo (funciona se estiver no PYZ ou sys.path correto)
    try:
        from . import sig_utils
    except (ImportError, ValueError):
        import sig_utils
    sign_pdf = sig_utils.sign_pdf
except Exception:
    # 2. Fallback absoluto se o import falhar (necessário em alguns casos do PyInstaller)
    import importlib.util
    _base_dir = os.path.dirname(os.path.abspath(__file__))
    _sig_utils_path = os.path.join(_base_dir, "sig_utils.py")
    
    # Failsafe para PyInstaller: verifica se o arquivo realmente existe fisicamente
    if not os.path.exists(_sig_utils_path):
        # 1. Tenta injetar 'PY' no caminho
        _alt_path = os.path.join(os.path.dirname(_base_dir), "PY", "assiname_app", "sig_utils.py")
        # 2. Tenta remover 'PY' do caminho se ele já estiver lá
        if not os.path.exists(_alt_path):
            if "\\PY\\" in _sig_utils_path: _alt_path = _sig_utils_path.replace("\\PY\\", "\\")
            elif "/PY/" in _sig_utils_path: _alt_path = _sig_utils_path.replace("/PY/", "/")
        
        if os.path.exists(_alt_path):
            _sig_utils_path = _alt_path
            
    _spec = importlib.util.spec_from_file_location("sig_utils", _sig_utils_path)
    _sig_utils = importlib.util.module_from_spec(_spec)
    _spec.loader.exec_module(_sig_utils)
    sign_pdf = _sig_utils.sign_pdf

app = Flask(__name__)
app.secret_key = "authetique_style_system_key_2026_hd"

# Diretorios e Caminhos - REDIRECIONADOS PARA LOCALAPPDATA PARA EVITAR ACESSO NEGADO NO DISCO C
def get_data_dir():
    data_root = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), "JUVICKS_DATA")
    # Se estiver em desenvolvimento (não frozen), pode usar a pasta local
    if not getattr(sys, 'frozen', False):
        return os.path.abspath(os.path.dirname(__file__))
    os.makedirs(data_root, exist_ok=True)
    return data_root

BASE_DIR = get_data_dir()
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
SIGNED_FOLDER = os.path.join(BASE_DIR, "signed")
DB_PATH = os.path.join(BASE_DIR, "assinaturas.db")

for f in [UPLOAD_FOLDER, SIGNED_FOLDER]: 
    os.makedirs(f, exist_ok=True)

# Pasta para orçamentos estáticos
ORC_FOLDER = UPLOAD_FOLDER
static_orc_dir = os.path.join(BASE_DIR, "static_orc")
os.makedirs(static_orc_dir, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SIGNED_FOLDER'] = SIGNED_FOLDER

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_base_url():
    """Tenta descobrir a URL pública do Ngrok para port 5001, ou usa host_url"""
    try:
        import requests
        # Aumentamos o timeout para 1.0s para ser mais robusto
        resp = requests.get("http://127.0.0.1:4040/api/tunnels", timeout=1.0)
        if resp.status_code == 200:
            for t in resp.json().get('tunnels', []):
                addr = t.get('config', {}).get('addr', '')
                # Verifica se é o túnel da porta 5001 (assiname/orcamentos)
                if ":5001" in addr or addr.endswith("5001"):
                    public_url = t.get('public_url')
                    if public_url:
                        return public_url.rstrip('/')
    except Exception as e:
        print(f"[DEBUG GLB] Erro ao detectar Ngrok: {e}")
    
    return request.host_url.rstrip('/')

def init_db():
    conn = get_db_connection()
    # Tabela de Documentos
    conn.execute('''
        CREATE TABLE IF NOT EXISTS documents (
            id TEXT PRIMARY KEY,
            original_filename TEXT NOT NULL,
            filename TEXT NOT NULL,
            whatsapp_number TEXT NOT NULL,
            status TEXT DEFAULT 'pending',
            created_at TEXT,
            client_ip TEXT,
            client_location TEXT,
            client_device_id TEXT,
            client_signature_time TEXT,
            sender_ip TEXT,
            sender_signature_time TEXT
        )
    ''')
    # Tabela de Orçamentos
    conn.execute('''
        CREATE TABLE IF NOT EXISTS orcamentos (
            id TEXT PRIMARY KEY,
            cliente_nome TEXT,
            cliente_tel TEXT,
            valor_total REAL,
            original_filename TEXT NOT NULL,
            filename TEXT NOT NULL,
            whatsapp_number TEXT,
            status TEXT DEFAULT 'pending',
            created_at TEXT,
            viewed_at TEXT,
            viewed_by TEXT
        )
    ''')
    # Tabela de Usuarios (Login Multiplo)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL, -- SHA-256
            role TEXT DEFAULT 'admin'
        )
    ''')
    import hashlib
    default_pass = hashlib.sha256("admin123".encode()).hexdigest()
    conn.execute("INSERT OR IGNORE INTO users (username, password) VALUES ('admin', ?)", (default_pass,))
    conn.commit()
    conn.close()

init_db()

# --- LOGIN / LOGOUT ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        import hashlib
        pass_hash = hashlib.sha256(password.encode()).hexdigest()
        
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE username = ? AND password = ?", (username, pass_hash)).fetchone()
        conn.close()
        
        if user:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Usuário ou senha inválidos', 'error')
    return render_template('login.html')

import requests
from flask import Response

@app.route('/')
def index():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

# --- FILTROS E PESQUISAS ---

@app.route('/')
@login_required
def admin_dashboard():
    tab = request.args.get('tab', 'envio')
    search = request.args.get('search', '')
    
    conn = get_db_connection()
    
    # Se estiver nas abas de visualizacao de dados, pegamos todos para preencher as colunas
    # A busca agora filtrara em todas as colunas simultaneamente
    query = "SELECT * FROM documents WHERE (original_filename LIKE ? OR whatsapp_number LIKE ?) ORDER BY created_at DESC"
    documents = conn.execute(query, (f'%{search}%', f'%{search}%')).fetchall()
    
    users = conn.execute("SELECT * FROM users").fetchall()
    conn.close()
    
    return render_template('admin.html', documents=documents, host_url=get_base_url(), active_tab=tab, search=search, users=users)

# --- UPLOAD E ENVIO ---

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'document' not in request.files:
        flash('Nenhum arquivo enviado', 'error')
        return redirect(url_for('admin_dashboard'))
    
    file = request.files['document']
    whatsapp = request.form.get('whatsapp', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    
    if file.filename == '' or not whatsapp:
        flash('Arquivo ou numero de WhatsApp ausente.', 'error')
        return redirect(url_for('admin_dashboard'))
        
    doc_id = str(uuid.uuid4())
    original_filename = secure_filename(file.filename)
    filename = f"{doc_id}_{original_filename}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    conn = get_db_connection()
    conn.execute('''
        INSERT INTO documents (id, original_filename, filename, whatsapp_number, created_at) 
        VALUES (?, ?, ?, ?, ?)
    ''', (doc_id, original_filename, filename, whatsapp, created_at))
    conn.commit()
    conn.close()
    
    flash('Documento pronto! Envie agora pelo botao do WhatsApp.', 'success')
    return redirect(url_for('admin_dashboard'))

# --- ORCAMENTOS ---

@app.route('/api/upload_pdf', methods=['POST'])
def upload_pdf_api():
    """Recebe PDF e retorna URL direta para download"""
    data = request.json
    nome_arquivo = data.get('nome_arquivo', 'documento.pdf')
    pdf_base64 = data.get('pdf_base64')
    
    if not pdf_base64:
        return jsonify({'erro': 'PDF não enviado'}), 400
    
    import base64
    pdf_bytes = base64.b64decode(pdf_base64)
    
    file_id = str(uuid.uuid4())[:8]
    filename = f"{file_id}_{nome_arquivo}"
    static_path = os.path.join(BASE_DIR, "static", "orc", filename)
    os.makedirs(os.path.dirname(static_path), exist_ok=True)
    
    with open(static_path, 'wb') as f:
        f.write(pdf_bytes)
    
    url = f"{get_base_url()}/static/orc/{filename}"
    return jsonify({'url': url})

@app.route('/api/enviar_orcamento', methods=['POST'])
def enviar_orcamento():
    """Recebe PDF do orçamento e gera link para visualização"""
    data = request.json
    
    orc_id = data.get('id')
    cliente_nome = data.get('cliente_nome', 'Cliente')
    cliente_tel = data.get('cliente_tel', '')
    valor_total = data.get('valor_total', 0)
    pdf_base64 = data.get('pdf_base64')
    
    if not pdf_base64:
        return jsonify({'erro': 'PDF não enviado'}), 400
    
    import base64
    pdf_bytes = base64.b64decode(pdf_base64)
    
    filename = f"Orc_{orc_id}.pdf"
    static_path = os.path.join(BASE_DIR, "static", "orc", filename)
    os.makedirs(os.path.dirname(static_path), exist_ok=True)
    
    with open(static_path, 'wb') as f:
        f.write(pdf_bytes)
    
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    conn = get_db_connection()
    conn.execute('''
        INSERT OR REPLACE INTO orcamentos (id, cliente_nome, cliente_tel, valor_total, original_filename, filename, whatsapp_number, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (orc_id, cliente_nome, cliente_tel, valor_total, filename, filename, cliente_tel, created_at))
    conn.commit()
    conn.close()
    
    link = f"{get_base_url()}/orc/{orc_id}"
    return jsonify({'sucesso': True, 'link': link})

@app.route('/orc/<orc_id>')
def ver_orcamento(orc_id):
    """Página de visualização do orçamento - PDF embutido em base64 para evitar dupla requisição ao Ngrok"""
    try:
        conn = get_db_connection()
        orc = conn.execute('SELECT * FROM orcamentos WHERE id = ?', (orc_id,)).fetchone()
        conn.close()

        if not orc:
            return f"Orçamento '{orc_id}' não encontrado no banco de dados.", 404

        # Ler o PDF do disco e converter para base64
        pdf_path = os.path.join(BASE_DIR, "static", "orc", orc['filename'])
        pdf_base64 = None
        if os.path.exists(pdf_path):
            import base64 as b64
            with open(pdf_path, 'rb') as f:
                pdf_base64 = b64.b64encode(f.read()).decode('utf-8')

        from flask import make_response
        renderizado = render_template(
            'orcamento_view.html',
            orc=orc,
            host_url=get_base_url(),
            pdf_base64=pdf_base64
        )
        response = make_response(renderizado)
        response.headers['ngrok-skip-browser-warning'] = 'true'
        return response
    except Exception as e:
        import traceback
        return f"<pre>ERRO: {str(e)}\n\n{traceback.format_exc()}</pre>", 500

@app.route('/static/orc/<filename_>')
def servir_orcamento(filename_):
    """Serve os PDFs dos orçamentos com headers robustos"""
    from flask import send_from_directory, make_response
    static_path = os.path.join(BASE_DIR, "static", "orc")
    
    response = make_response(send_from_directory(static_path, filename_))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=' + filename_
    # Bypass para aviso do Ngrok se o cliente for um visualizador ou browser
    response.headers['ngrok-skip-browser-warning'] = 'true'
    return response

@app.route('/api/marcar_visto/<orc_id>', methods=['POST'])
def marcar_orcamento_visto(orc_id):
    """Marca que o orçamento foi visualizado"""
    viewed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    viewed_by = request.remote_addr
    
    conn = get_db_connection()
    conn.execute('UPDATE orcamentos SET viewed_at = ?, viewed_by = ? WHERE id = ?', (viewed_at, viewed_by, orc_id))
    conn.commit()
    conn.close()
    
    return jsonify({'sucesso': True})

# --- VISUALIZACAO CLIENTE ---

@app.route('/sign/<doc_id>')
def client_view(doc_id):
    conn = get_db_connection()
    doc = conn.execute('SELECT * FROM documents WHERE id = ?', (doc_id,)).fetchone()
    conn.close()
    
    if not doc:
        return render_template('error.html', message="Documento nao encontrado."), 404
    
    # Se ja estiver assinado ou rejeitado, apenas mostrar o PDF final
    return render_template('client_view.html', doc=doc)

@app.route('/api/sign_client/<doc_id>', methods=['POST'])
def sign_action_client(doc_id):
    # Capturar informacoes em JSON (Location e Device ID)
    data = request.json
    location = data.get('location', 'Nao autorizado')
    device_id = data.get('device_id', 'N/A')
    visual_signature = data.get('visual_signature') # Base64 do Canvas ou Imagem
    client_ip = request.remote_addr
    
    conn = get_db_connection()
    doc = conn.execute('SELECT * FROM documents WHERE id = ?', (doc_id,)).fetchone()
    
    if not doc or doc['status'] != 'pending':
        conn.close()
        return jsonify({"success": False, "message": "Documento ja processado ou nao encontrado."}), 400

    input_pdf = os.path.join(app.config['UPLOAD_FOLDER'], doc['filename'])
    output_pdf = os.path.join(app.config['SIGNED_FOLDER'], f"client_{doc['filename']}")
    
    sign_data = {
        "name": f"Cliente ({doc['whatsapp_number']})",
        "id": doc_id,
        "original_filename": doc['original_filename'],
        "ip": client_ip,
        "location": location,
        "mac_address": data.get('mac_address', 'UID-UNDEF'),
        "visual_signature": visual_signature
    }
    
    try:
        sign_pdf(input_pdf, output_pdf, sign_data, "CLIENT")
        
        signature_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute('''
            UPDATE documents 
            SET status = 'client_signed', client_ip = ?, client_location = ?, 
                client_device_id = ?, client_signature_time = ? 
            WHERE id = ?
        ''', (client_ip, location, device_id, signature_time, doc_id))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": f"Erro no PDF: {str(e)}"}), 500

@app.route('/action/reject/<doc_id>', methods=['POST'])
def action_reject(doc_id):
    conn = get_db_connection()
    conn.execute("UPDATE documents SET status = 'rejected' WHERE id = ?", (doc_id,))
    conn.commit()
    conn.close()
    flash('Documento marcado como recusado pelo cliente.', 'info')
    return redirect(url_for('client_view', doc_id=doc_id))

@app.route('/action/delete/<doc_id>', methods=['POST'])
@login_required
def action_delete(doc_id):
    doc_id = str(doc_id)
    conn = get_db_connection()
    doc = conn.execute('SELECT filename FROM documents WHERE id = ?', (doc_id,)).fetchone()
    
    if doc:
        fname = doc['filename']
        # Tenta apagar todos os arquivos possiveis vinculados
        paths = [
            os.path.join(app.config['UPLOAD_FOLDER'], fname),
            os.path.join(app.config['SIGNED_FOLDER'], f"client_{fname}"),
            os.path.join(app.config['SIGNED_FOLDER'], f"completed_{fname}")
        ]
        for p in paths:
            try:
                if os.path.exists(p): os.remove(p)
            except: pass

        conn.execute('DELETE FROM documents WHERE id = ?', (doc_id,))
        conn.commit()
        flash('Documento e arquivos removidos com sucesso!', 'success')
    
    conn.close()
    return redirect(request.referrer or url_for('admin_dashboard'))

# --- ACESSO ADMIN ---

@app.route('/action/sign_admin/<doc_id>', methods=['POST'])
@login_required
def action_sign_admin(doc_id):
    data = request.json
    visual_signature = data.get('visual_signature') # Base64
    
    conn = get_db_connection()
    doc = conn.execute('SELECT * FROM documents WHERE id = ?', (doc_id,)).fetchone()
    
    if not doc or doc['status'] != 'client_signed':
        conn.close()
        return jsonify({"success": False, "message": "Apenas documentos assinados pelo cliente podem ser finalizados."})
    
    input_pdf = os.path.join(app.config['SIGNED_FOLDER'], f"client_{doc['filename']}")
    output_pdf = os.path.join(app.config['SIGNED_FOLDER'], f"completed_{doc['filename']}")
    
    sign_data = {
        "name": "Administrador / Empresa",
        "id": doc_id,
        "original_filename": doc['original_filename'],
        "ip": request.remote_addr,
        "location": "Sede Administrativa",
        "mac_address": f"ADM-{os.name.upper()}-{os.cpu_count() or 0}",
        "visual_signature": visual_signature
    }
    
    try:
        sign_pdf(input_pdf, output_pdf, sign_data, "SENDER")
        signature_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute('''
            UPDATE documents 
            SET status = 'completed', sender_ip = ?, sender_signature_time = ? 
            WHERE id = ?
        ''', (request.remote_addr, signature_time, doc_id))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})

# --- IMPORT / EXPORT DB ---

@app.route('/export')
@login_required
def export_data():
    conn = get_db_connection()
    cursor = conn.execute("SELECT * FROM documents")
    rows = cursor.fetchall()
    columns = [description[0] for description in cursor.description]
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assinaturas"

    # Header
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Data
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    export_path = os.path.join(app.config['UPLOAD_FOLDER'], 'assinaturas_backup.xlsx')
    wb.save(export_path)
    return send_file(export_path, as_attachment=True)

@app.route('/import', methods=['POST'])
@login_required
def import_data():
    if 'file' not in request.files: return redirect(url_for('admin_dashboard', tab='backup'))
    file = request.files['file']
    if file.filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        conn = get_db_connection()
        for i in range(2, ws.max_row + 1):
            row_values = []
            for j in range(1, ws.max_column + 1):
                row_values.append(ws.cell(row=i, column=j).value)
            if not row_values[0]: continue
            placeholders = ', '.join(['?'] * len(row_values))
            try:
                conn.execute(f"INSERT OR REPLACE INTO documents ({', '.join(headers)}) VALUES ({placeholders})", tuple(row_values))
            except: pass
        conn.commit()
        conn.close()
        flash('Excel importado!', 'success')
    return redirect(url_for('admin_dashboard', tab='backup'))

# --- CONFIGURACAO DE USUARIOS ---

@app.route('/settings/user/add', methods=['POST'])
@login_required
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    import hashlib
    pass_hash = hashlib.sha256(password.encode()).hexdigest()
    
    conn = get_db_connection()
    try:
        conn.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, pass_hash))
        conn.commit()
        flash(f'Usuario {username} criado!', 'success')
    except:
        flash('Erro: Usuario ja existe.', 'error')
    conn.close()
    return redirect(url_for('admin_dashboard', tab='config'))

@app.route('/settings/user/delete/<username>', methods=['POST'])
@login_required
def delete_user(username):
    if username == 'admin':
        flash('Nao eh possivel apagar o admin principal.', 'error')
        return redirect(url_for('admin_dashboard', tab='config'))
    
    conn = get_db_connection()
    conn.execute("DELETE FROM users WHERE username = ?", (username,))
    conn.commit()
    conn.close()
    flash(f'Usuario {username} removido!', 'success')
    return redirect(url_for('admin_dashboard', tab='config'))

# --- BACKUP .DB ---

@app.route('/backup/export_db')
@login_required
def export_db():
    return send_file(DB_PATH, as_attachment=True, download_name=f"backup_assinahd_{datetime.now().strftime('%d%m%Y')}.db")

@app.route('/backup/import_db', methods=['POST'])
@login_required
def import_db():
    if 'file' not in request.files: return redirect(url_for('admin_dashboard', tab='backup'))
    file = request.files['file']
    if file.filename.endswith('.db'):
        file.save(DB_PATH) # Cuidado: isso sobrescreve o atual!
        flash('Banco de Dados (.db) restaurado com sucesso!', 'success')
    return redirect(url_for('admin_dashboard', tab='backup'))

@app.route('/download/<doc_id>')
def download_pdf(doc_id):
    conn = get_db_connection()
    doc = conn.execute('SELECT * FROM documents WHERE id = ?', (doc_id,)).fetchone()
    conn.close()
    if not doc: return "404", 404
    
    if doc['status'] == 'completed':
        path = os.path.join(app.config['SIGNED_FOLDER'], f"completed_{doc['filename']}")
    elif doc['status'] == 'client_signed':
        path = os.path.join(app.config['SIGNED_FOLDER'], f"client_{doc['filename']}")
    else:
        path = os.path.join(app.config['UPLOAD_FOLDER'], doc['filename'])
        
    if os.path.exists(path): return send_file(path)
    return "File not found", 404

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--port', type=int, default=int(os.environ.get('FLASK_PORT', 5001)))
    args, _ = parser.parse_known_args()
    print(f"[FLASK] Iniciando na porta {args.port}...")
    app.run(debug=False, host='0.0.0.0', port=args.port)
